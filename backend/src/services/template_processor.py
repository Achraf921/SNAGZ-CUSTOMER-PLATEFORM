from openpyxl import load_workbook
import json
import sys
import os
import base64
import datetime

# Set up logging to a file
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f'template_processor_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    def log_message(message):
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f'[{timestamp}] {message}\n')
            print(f'[{timestamp}] {message}')
    
    return log_message

log = setup_logging()

def process_template_xlsx(template_path, template_data, output_path):
    log(f"Starting Template D2C processing for: {template_path}")
    log(f"Output will be saved to: {output_path}")
    
    try:
        # Load the template workbook
        workbook = load_workbook(template_path)
        log(f"Successfully loaded template: {template_path}")
        
        # Get the first worksheet
        worksheet = workbook.active
        log(f"Processing worksheet: {worksheet.title}")
        
        # Fill row 2 with template data according to column structure:
        # Col 1: nom de projet
        # Col 2: type de projet  
        # Col 3: commercial
        # Col 4: boutique en ligne
        # Col 5: client
        # Col 6: contacts client
        # Col 7: numero compte client
        # Col 8: date de mise en ligne
        # Col 9: date de commercialisation
        # Col 10: date de sortie officielle
        # Col 11: precommande (OUI/NON)
        # Col 12: dedicace (OUI/NON)
        # Col 13: facturation (mandataire/vendeur)
        # Col 14: abonnement mensuel shopify (OUI/NON)
        # Col 15: abonnement annuel shopify (OUI/NON)
        # Col 16: couts mondial relay (OUI/NON)
        # Col 17: couts delivengo (OUI/NON)
        # Col 18: frais mensuel maintenance (50€)
        # Col 19: frais ouverture boutique (500€)
        # Col 20: frais ouverture sans habillage (empty)
        # Col 21: commission snagz (pourcentageSNA%)
        
        row_data = [
            template_data.get('nomProjet', ''),                # Col 1
            template_data.get('typeProjet', ''),               # Col 2  
            template_data.get('commercial', ''),               # Col 3
            template_data.get('boutiqueEnLigne', ''),          # Col 4
            template_data.get('client', ''),                   # Col 5
            template_data.get('contactsClient', ''),           # Col 6
            template_data.get('numeroCompteClient', ''),       # Col 7
            template_data.get('dateMiseEnLigne', ''),          # Col 8
            template_data.get('dateCommercialisation', ''),    # Col 9
            template_data.get('dateSortieOfficielle', ''),     # Col 10
            template_data.get('precommande', ''),              # Col 11
            template_data.get('dedicace', ''),                 # Col 12
            template_data.get('facturation', ''),              # Col 13
            template_data.get('abonnementMensuelShopify', ''), # Col 14
            template_data.get('abonnementAnnuelShopify', ''),  # Col 15
            template_data.get('coutsMondialRelay', ''),        # Col 16
            template_data.get('coutsDelivengo', ''),           # Col 17
            template_data.get('fraisMensuelMaintenance', ''),  # Col 18
            template_data.get('fraisOuvertureBoutique', ''),   # Col 19
            template_data.get('fraisOuvertureSansHabillage', ''), # Col 20
            template_data.get('commissionSnagz', '')           # Col 21
        ]
        
        log(f"Filling row 2 with {len(row_data)} columns of data")
        
        # Fill row 2 (index 2) with the data
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.value = value
            log(f"Set column {col_idx} to: '{value}'")
        
        # Save the workbook
        workbook.save(output_path)
        log(f"Template D2C file saved to {output_path}")
        log("Template D2C processing completed successfully")
        
    except Exception as e:
        log(f"Error processing template: {str(e)}")
        raise e

def main():
    if len(sys.argv) != 4:
        print("Usage: python3 template_processor.py <template_path> <encoded_data> <output_path>")
        sys.exit(1)
    
    template_path = sys.argv[1]
    encoded_data = sys.argv[2]
    output_path = sys.argv[3]
    
    log("Template processor script started")
    log(f"Template path: {template_path}")
    log(f"Output path: {output_path}")
    log(f"Encoded data length: {len(encoded_data)} characters")
    
    try:
        # Decode the base64 data
        log("Decoding base64 data...")
        decoded_data = base64.b64decode(encoded_data).decode('utf-8')
        log("Successfully decoded base64 data")
        
        # Parse JSON data
        log("Parsing JSON data...")
        template_data = json.loads(decoded_data)
        log("Successfully parsed JSON data")
        log(f"Template data keys: {list(template_data.keys())}")
        
        # Process the template
        log("Starting template processing...")
        process_template_xlsx(template_path, template_data, output_path)
        
    except Exception as e:
        log(f"Fatal error: {str(e)}")
        print(f"Fatal error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main() 