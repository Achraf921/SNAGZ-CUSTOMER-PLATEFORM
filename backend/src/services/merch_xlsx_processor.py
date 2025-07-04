from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side
import json
import sys
import os
import base64
import datetime

# Set up logging to a file
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f'merch_xlsx_processor_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    def log_message(message):
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f'[{timestamp}] {message}\n')
            print(f'[{timestamp}] {message}')
    
    return log_message

log = setup_logging()

def calculate_total_stock(stock_dict):
    """Calculate total stock from stock dictionary"""
    if not stock_dict:
        return 0
    return sum(int(value) for value in stock_dict.values() if str(value).isdigit())

def calculate_size_stock(stock_dict, target_size):
    """Calculate stock for a specific size regardless of color"""
    if not stock_dict:
        return 0
    
    total = 0
    for combination, stock in stock_dict.items():
        if '-' in combination:
            size, color = combination.split('-', 1)
            if size == target_size and str(stock).isdigit():
                total += int(stock)
        elif combination == target_size and str(stock).isdigit():
            total += int(stock)
    
    return total

def process_merch_xlsx(xlsx_path, shop_data, output_path):
    log(f"Starting Merchandising XLSX processing for: {xlsx_path}")
    log(f"Output will be saved to: {output_path}")
    
    try:
        workbook = load_workbook(xlsx_path)
        log(f"Successfully loaded workbook: {xlsx_path}")
    except Exception as e:
        log(f"Failed to load workbook: {str(e)}")
        raise

    log(f"Shop data keys: {list(shop_data.keys())}")
    
    # Extract shop info and products
    nom_projet = shop_data.get('nomProjet', '')
    shopify_domain = shop_data.get('shopifyDomain', '')
    products = shop_data.get('products', [])
    
    # Debug: Check for date-related fields
    for key in shop_data.keys():
        if 'date' in key.lower() or 'sortie' in key.lower() or 'commercialisation' in key.lower():
            log(f"Found date-related field: '{key}' = '{shop_data[key]}'")
    
    # Debug: Check for customer-related fields
    for key in shop_data.keys():
        if 'raison' in key.lower() or 'customer' in key.lower() or 'client' in key.lower():
            log(f"Found customer-related field: '{key}' = '{shop_data[key]}'")
    
    # Extract shop-level data for additional fields - try multiple field name variations
    date_sortie = (shop_data.get('dateSortie', '') or 
                   shop_data.get('dateDeSortie', '') or 
                   shop_data.get('dateAlbum', '') or 
                   shop_data.get('dateSortieAlbum', ''))
    
    date_commercialisation = (shop_data.get('dateCommercialisation', '') or 
                             shop_data.get('dateDeCommercialisation', '') or 
                             shop_data.get('dateMerch', '') or 
                             shop_data.get('dateCommercialisationMerch', ''))
    
    raison_sociale = shop_data.get('raisonSociale', '') or shop_data.get('customerName', '')
    
    log(f"Extracted dates - Sortie: '{date_sortie}', Commercialisation: '{date_commercialisation}'")
    log(f"Extracted raison sociale: '{raison_sociale}'")
    
    # Combine dates for column 14 (Date de sortie + Date commercialisation)
    combined_dates = ''
    if date_sortie:
        combined_dates += f"DATE DE SORTIE (Album): {date_sortie}"
    if date_commercialisation:
        if combined_dates:
            combined_dates += f"\nCOMMERCIALISATION (Merch): {date_commercialisation}"
        else:
            combined_dates = f"COMMERCIALISATION (Merch): {date_commercialisation}"
    
    log(f"Combined dates field: '{combined_dates}'")
    
    log(f"Processing {len(products)} products for shop: {nom_projet}")

    # Process all worksheets
    for worksheet in workbook.worksheets:
        log(f"Processing worksheet: {worksheet.title}")
        
        # Replace placeholders in header, but skip merged cells
        for row in worksheet.iter_rows():
            for cell in row:
                # Skip merged cells - they can't be modified directly
                if isinstance(cell, MergedCell):
                    continue
                    
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    new_value = original_value
                    
                    # Replace placeholders - try multiple variations
                    placeholders_to_replace = [
                        ('nonProjet', nom_projet),
                        ('nomProjet', nom_projet),
                        ('CLIENT', nom_projet),
                        ('PROJET', nom_projet),
                        ('shopifyDomain', shopify_domain),
                        ('SHOPIFY_DOMAIN', shopify_domain)
                    ]
                    
                    for placeholder, replacement in placeholders_to_replace:
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, replacement)
                            log(f"Replaced '{placeholder}' with '{replacement}' in cell {cell.coordinate}")
                    
                    if new_value != original_value:
                        try:
                            cell.value = new_value
                        except AttributeError as e:
                            log(f"Could not modify cell {cell.coordinate}: {str(e)}")
                            continue

        # Find the row that contains size headers (XS, S, M, L, XL)
        # This is likely the last header row before data should be written
        data_start_row = None
        size_header_row = None
        
        # First, look for the row with size headers (XS, S, M, L, XL)
        for row_num in range(1, worksheet.max_row + 15):  # Check more rows
            found_sizes = 0
            size_positions = {}
            
            # Check all columns in this row for size headers
            for col_num in range(1, 15):  # Check first 15 columns
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value and isinstance(cell_value, str):
                    cell_value_upper = cell_value.strip().upper()
                    if cell_value_upper in ['XS', 'S', 'M', 'L', 'XL']:
                        found_sizes += 1
                        size_positions[cell_value_upper] = col_num
            
            # If we found at least 3 size headers, this is likely the size header row
            if found_sizes >= 3:
                size_header_row = row_num
                data_start_row = row_num + 1  # Start writing data AFTER the size headers
                log(f"Found size header row at row {size_header_row} with {found_sizes} size columns: {size_positions}")
                break
        
        # Fallback: Look for "TYPE DE PRODUIT" in the first cell
        if data_start_row is None:
            for row_num in range(1, worksheet.max_row + 15):
                first_cell_value = worksheet.cell(row=row_num, column=1).value
                if first_cell_value and isinstance(first_cell_value, str):
                    if "TYPE DE PRODUIT" in first_cell_value.upper():
                        # Look for size headers in the next few rows after TYPE DE PRODUIT
                        for next_row in range(row_num + 1, min(row_num + 5, worksheet.max_row + 1)):
                            found_sizes = 0
                            for col_num in range(1, 15):
                                cell_value = worksheet.cell(row=next_row, column=col_num).value
                                if cell_value and isinstance(cell_value, str):
                                    if cell_value.strip().upper() in ['XS', 'S', 'M', 'L', 'XL']:
                                        found_sizes += 1
                            if found_sizes >= 3:
                                size_header_row = next_row
                                data_start_row = next_row + 1
                                log(f"Found 'TYPE DE PRODUIT' at row {row_num}, size headers at row {size_header_row}")
                                break
                        if data_start_row:
                            break
                        else:
                            # No size headers found, use row after TYPE DE PRODUIT + buffer
                            data_start_row = row_num + 3
                            log(f"Found 'TYPE DE PRODUIT' at row {row_num}, using row {data_start_row} with buffer")
                            break
        
        # Second fallback: look for any cell containing "TYPE" in the first column
        if data_start_row is None:
            for row_num in range(1, worksheet.max_row + 15):
                first_cell_value = worksheet.cell(row=row_num, column=1).value
                if first_cell_value and isinstance(first_cell_value, str):
                    if "TYPE" in first_cell_value.upper():
                        data_start_row = row_num + 3  # Add buffer to avoid overwriting
                        log(f"Found header row with 'TYPE' at row {row_num}, using row {data_start_row} with buffer")
                        break
        
        # Final fallback
        if data_start_row is None:
            data_start_row = 6  # Start at row 6 to be safe
            log("Could not find header rows - using default row 6 with safety buffer")
            
        log(f"Will start inserting product data at row: {data_start_row}")
        
        # Add products data
        current_row = data_start_row
        log(f"Starting to add {len(products)} products at row {current_row}")
        
        for i, product in enumerate(products):
            log(f"Processing product {i+1}/{len(products)}: {product.get('titre', 'Unknown')}")
            
            # Calculate stock values
            stock_dict = product.get('stock', {})
            total_stock = calculate_total_stock(stock_dict)
            log(f"Product {i+1} total stock: {total_stock} from stock_dict: {stock_dict}")
            
            # Calculate stock by size
            size_stocks = {}
            for size in ['XS', 'S', 'M', 'L', 'XL']:
                size_stocks[size] = calculate_size_stock(stock_dict, size)
            log(f"Product {i+1} size stocks: {size_stocks}")
            
            # Row data according to specifications:
            # 1. Type de produit
            # 2. Titre  
            # 3. Description fiche produits
            # 4. Code-barres EAN (13 chiffres)
            # 5. Quantités totales
            # 6. Merch/tailles (XS, S, M, L, XL sub-columns)
            # 7. Poids
            # 8. Prix de vente TTC
            # 9. OCC (OUI/NON)
            
            # Handle different field name variations
            type_produit = product.get('typeProduit', '') or product.get('type', '')
            titre = product.get('titre', '') or product.get('title', '')
            description = product.get('description', '')
            code_ean = product.get('codeEAN', '') or product.get('ean', '') or product.get('codeBarres', '')
            poids = product.get('poids', '') or product.get('weight', '')
            prix = product.get('prix', '') or product.get('price', '')
            occ = product.get('occ', False) or product.get('OCC', False)
            
            # Extract colors and sizes from product
            couleurs = product.get('couleurs', [])
            tailles = product.get('tailles', [])
            
            # Format colors and sizes as comma-separated strings
            couleurs_str = ', '.join(couleurs) if isinstance(couleurs, list) else str(couleurs) if couleurs else ''
            tailles_str = ', '.join(tailles) if isinstance(tailles, list) else str(tailles) if tailles else ''
            
            log(f"Product {i+1} colors: {couleurs} -> '{couleurs_str}'")
            log(f"Product {i+1} sizes: {tailles} -> '{tailles_str}'")
            
            row_data = [
                type_produit,                            # Column 1: Type de produit
                titre,                                   # Column 2: Titre
                description,                             # Column 3: Description
                code_ean,                                # Column 4: Code EAN
                total_stock,                             # Column 5: Quantités totales (per product)
                size_stocks.get('XS', 0),               # Column 6a: XS stock
                size_stocks.get('S', 0),                # Column 6b: S stock
                size_stocks.get('M', 0),                # Column 6c: M stock
                size_stocks.get('L', 0),                # Column 6d: L stock
                size_stocks.get('XL', 0),               # Column 6e: XL stock
                poids,                                   # Column 7: Poids
                prix,                                    # Column 8: Prix TTC
                'OUI' if occ else 'NON',                # Column 13: OCC
                combined_dates,                          # Column 14: Date de sortie + Date commercialisation
                raison_sociale,                          # Column 15: Artiste/Label (Raison sociale from customer)
                '',                                      # Column 16: Empty cell
                '',                                      # Column 17: Empty cell
                '',                                      # Column 18: Empty cell
                couleurs_str,                            # Column 19: Couleurs (colors selected by users)
                tailles_str,                             # Column 20: Tailles (sizes selected by users)
                ''                                       # Column 21: Visuels (empty cell)
            ]
            
            # Insert the row with borders
            for col_num, value in enumerate(row_data, 1):
                try:
                    cell = worksheet.cell(row=current_row, column=col_num)
                    # Skip if this is a merged cell
                    if isinstance(cell, MergedCell):
                        log(f"Skipping merged cell at row {current_row}, column {col_num}")
                        continue
                    
                    # Set cell value
                    cell.value = value
                    
                    # Add solid borders to all sides
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                    
                except Exception as e:
                    log(f"Error setting value for cell ({current_row}, {col_num}): {str(e)}")
                    continue
            
            log(f"Successfully added product {i+1} data at row {current_row}: {product.get('titre', 'Unknown')}")
            log(f"Combined dates field content: '{combined_dates}'")
            log(f"Colors field content: '{couleurs_str}'")
            log(f"Sizes field content: '{tailles_str}'")
            log(f"Data written: {row_data}")
            current_row += 1
            
        log(f"Finished processing all products. Final row: {current_row - 1}")

    # Save the processed workbook
    workbook.save(output_path)
    log(f"Merchandising XLSX file saved to {output_path}")

if __name__ == "__main__":
    try:
        log("Merchandising XLSX processor script started")
        
        if len(sys.argv) != 4:
            error_msg = "Usage: python merch_xlsx_processor.py <xlsx_template_path> <shop_data_json_string> <output_xlsx_path>"
            log(error_msg)
            print(error_msg, file=sys.stderr)
            sys.exit(1)

        template_path = sys.argv[1]
        encoded_shop_data_string = sys.argv[2]
        output_path = sys.argv[3]

        log(f"Template path: {template_path}")
        log(f"Output path: {output_path}")

        try:
            # Decode the Base64 string back to JSON
            log("Decoding base64 data...")
            shop_data_string = base64.b64decode(encoded_shop_data_string).decode('utf-8')
            log("Successfully decoded base64 data")

            # Parse the JSON string
            log("Parsing JSON data...")
            shop_data = json.loads(shop_data_string)
            log("Successfully parsed JSON data")

            # Process the XLSX file
            log("Starting merchandising XLSX processing...")
            process_merch_xlsx(template_path, shop_data, output_path)
            log("Merchandising XLSX processing completed successfully")
            
        except Exception as e:
            error_msg = f"Error processing merchandising XLSX: {str(e)}"
            log(error_msg)
            raise
            
    except Exception as e:
        error_msg = f"Fatal error: {str(e)}"
        log(error_msg)
        print(error_msg, file=sys.stderr)
        sys.exit(1) 