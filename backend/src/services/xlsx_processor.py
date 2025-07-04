from openpyxl import load_workbook
import json
import sys
import os
import base64
import datetime

# Set up logging to a file
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f'xlsx_processor_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    def log_message(message):
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f'[{timestamp}] {message}\n')
            print(f'[{timestamp}] {message}')
    
    return log_message

log = setup_logging()


def replace_placeholders_in_xlsx(xlsx_path, shop_data, output_path):
    log(f"Starting XLSX processing for: {xlsx_path}")
    log(f"Output will be saved to: {output_path}")
    
    try:
        workbook = load_workbook(xlsx_path)
        log(f"Successfully loaded workbook: {xlsx_path}")
    except Exception as e:
        log(f"Failed to load workbook: {str(e)}")
        raise

    log(f"[DEBUG - Inside replace_placeholders_in_xlsx] Type of shop_data: {type(shop_data)}")
    log(f"[DEBUG - Inside replace_placeholders_in_xlsx] Content of shop_data: {shop_data}")
    
    # Ensure shop_data is a dictionary
    if not isinstance(shop_data, dict):
        error_msg = f"[ERROR] shop_data is not a dictionary. Type: {type(shop_data)}, Value: {shop_data}"
        log(error_msg)
        raise ValueError(error_msg)

    # Get the raisonSociale from the customer data (passed in shop_data)
    raison_sociale = shop_data.get("raisonSociale", "") or shop_data.get("clientName", "").split('_')[0]
    
    # Get boolean values from shop_data and convert to OUI/NON
    precommande_value = "OUI" if shop_data.get("precommande", False) else "NON"
    dedicace_value = "OUI" if shop_data.get("dedicaceEnvisagee", False) else "NON"
    
    placeholder_mapping = {
        "XXX1": shop_data.get("nomProjet", ""),  # Project name
        "XXX2": shop_data.get("typeProjet", ""),  # Project type
        "XXX3": shop_data.get("commercial", ""),  # XXX3 commercial
        "XXX4": raison_sociale,  # Client's raison sociale
        "XXX5": shop_data.get("compteClientRef", ""),  # Client reference
        "XXX6": shop_data.get("contactsClient", ""),  # Client contact email
        "XXX7": shop_data.get("dateMiseEnLigne", ""),  # Go-live date
        "XXX8": shop_data.get("dateCommercialisation", ""),  # Commercialization date
        "XXX9": shop_data.get("dateSortieOfficielle", ""),  # Official release date
        "XXX10": precommande_value,  # Pre-order (OUI/NON)
        "XXX11": dedicace_value,  # Dedication planned (OUI/NON)
        "XXX12": shop_data.get("boutiqueEnLigne", ""),  # Boutique en ligne (OUI/NON)
        "XXX13": shop_data.get("chefProjet", ""),  # Chef de projet (prenom + nom)
        "XXX14": shop_data.get("demarrageProjet", ""),
        "XXX15": shop_data.get("contactsClient", ""),
        "COMPTENUM": shop_data.get("compteClientRef", ""),  # Client reference number
    }

    log(f"Processing {len(workbook.worksheets)} worksheets...")

    # Process all worksheets
    for worksheet in workbook.worksheets:
        log(f"Processing worksheet: {worksheet.title}")
        
        # Iterate through all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    new_value = original_value
                    
                    # Replace all placeholders in the cell value
                    for placeholder, replacement in placeholder_mapping.items():
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, str(replacement))
                            log(f"Replaced {placeholder} with {replacement} in cell {cell.coordinate}")
                    
                    # Update the cell if any replacements were made
                    if new_value != original_value:
                        cell.value = new_value

    # Save the processed workbook
    workbook.save(output_path)
    log(f"XLSX file saved to {output_path}")

if __name__ == "__main__":
    try:
        log("XLSX processor script started")
        
        if len(sys.argv) != 4:
            error_msg = "Usage: python xlsx_processor.py <xlsx_template_path> <shop_data_json_string> <output_xlsx_path>"
            log(error_msg)
            print(error_msg, file=sys.stderr)
            sys.exit(1)

        template_path = sys.argv[1]
        encoded_shop_data_string = sys.argv[2]
        output_path = sys.argv[3]

        log(f"Template path: {template_path}")
        log(f"Output path: {output_path}")
        log(f"Encoded data length: {len(encoded_shop_data_string)} characters")

        try:
            # Decode the Base64 string back to JSON
            log("Decoding base64 data...")
            shop_data_string = base64.b64decode(encoded_shop_data_string).decode('utf-8')
            log("Successfully decoded base64 data")
            
            log(f"[DEBUG] Type of shop_data_string after decode: {type(shop_data_string)}")
            log(f"[DEBUG] First 100 chars of shop_data_string: {shop_data_string[:100]}...")

            # Parse the JSON string
            log("Parsing JSON data...")
            shop_data = json.loads(shop_data_string)
            log("Successfully parsed JSON data")
            
            log(f"[DEBUG] Type of shop_data after json.loads: {type(shop_data)}")
            log(f"[DEBUG] shop_data keys: {list(shop_data.keys())}")

            # Process the XLSX file
            log("Starting XLSX processing...")
            replace_placeholders_in_xlsx(template_path, shop_data, output_path)
            log("XLSX processing completed successfully")
            
        except Exception as e:
            error_msg = f"Error processing XLSX: {str(e)}"
            log(error_msg)
            log(f"Error type: {type(e).__name__}")
            if hasattr(e, 'args'):
                log(f"Error args: {e.args}")
            raise
            
    except Exception as e:
        error_msg = f"Fatal error: {str(e)}"
        log(error_msg)
        print(error_msg, file=sys.stderr)
        sys.exit(1) 